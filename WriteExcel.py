from Report import *
from openpyxl import *
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import random

def writeToExcel():
    """Variables """
    dictlist = [assigndict,quizdict,labdict,otherdict,examdict]
    letters = 'ABCDEFGHIJK'  # to add more
    l_count = 0
    num_count = 8
    max_letter = 'A',
    max_num = 0

    # Algorithm to add name and grade to excel sheet
    for dict in dictlist:
        if len(dict) > 1:  # if dictionary isn't empty

            for key in dict:  # iterating through each dictionary
                """NAMES"""
                aName = ws[str(letters[l_count]) + str(num_count)]
                aName.value = key
                if num_count != 8:
                    aName.font = Font(bold=True)
                '''VALUES'''
                ws[str(letters[l_count + 1]) + str(num_count)] = dict[key]
                num_count += 1

            max_letter = letters[l_count + 1]
            """FINDING THE LAST CELL"""
            if num_count > max_num:
                max_num = num_count
            '''Row count resets back to 8
                skips to two columns over'''
            num_count = 8
            l_count += 2

    return max_letter + str(max_num-1)

"""Iterating data into Excel table"""
try:
    true = True
    while true:
        exsting_or_new = input("Are you adding to a existing Excel workbook (y/n): ")
        if exsting_or_new.lower() == "y":
            try:
                file = input("Enter filename (not including '.xlsx'): ")
                wb = load_workbook(file + ".xlsx")
                true = False
                ws = wb.create_sheet(className)
            except FileNotFoundError:
                print("Sorry unable to find the Excel file you have entered. Please try again")

        elif exsting_or_new.lower() == "n":
            file = input("What name would you like to save your workbook to? (not including '.xlsx'): ")
            wb = Workbook()
            ws = wb.active
            true = False
        else:
            print("Invalid entry, please try again.")

    # creating excel sheet
    # modifying existing workbook, to create workbook its different function
    # how to start sheet
    ws.title = className  # names sheet

    # names the Grade Report title
    ws.merge_cells('A1:F5')
    title = ws["A1"]
    title.value = "Grade Report for " + className
    title.font = Font(size=15,italic=True,bold=True, underline='single')

    last_cell = writeToExcel()

    try:
        randomNum = random.randint(0, 10000)
        tab = Table(displayName="GradesTablev" + str(randomNum),
                    ref="A8:" + last_cell)  # calling header value calls last value form makingHeaders()
    except ValueError:
        randomNum = random.randint(0, 10000)
        tab = Table(displayName="GradesTablev" + str(randomNum),
                    ref="A8:" + last_cell)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(file + ".xlsx")

    print("\nTo autofit width of table in Microsoft Excel, highlight all of the gradeTable, "
      "click the FORMAT button on the top right, then click AUTOFIT COLUMN WIDTH.\n Enjoy!")


except PermissionError:
    print("Unable to convert to Excel file. \nMake sure the workbook is closed and try again.")

except FileNotFoundError:
    print("Sorry unable to find the Excel file you have entered. Please try again")

except ValueError:
    print("Code Error. Please try again.")
